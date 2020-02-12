#  EXCEL FEE-HELP SUMMARY REPORT #
#  ANGUS CLARKE FEB 2020 Ver 2.11 #

import tkinter
from tkinter import filedialog
import openpyxl
from os import chdir


# ____________TKINTER INITIALISATIONS_____________ #

window = tkinter.Tk(className="spreadsheet manager")
window.geometry("660x600")
window.filename = 'Empty'
window.directory = 'Empty'
buttonVal= tkinter.IntVar()
buttonVal.set(0)
buttonVal2 = tkinter.IntVar()
buttonVal2.set(0)
spreadsheets = [
    ('Fees and Enrolments', 0),
    ('Donations', 1),
]


#Gets button value
def showChoice():
    print(buttonVal.get())


def showChoice2():
    buttonVal2.set(1)
    print(buttonVal2.get())
    openFiles()


def output2():
    window.directory = filedialog.askdirectory(mustexist=True)
    tkinter.Label(window, text=window.directory, justify=tkinter.LEFT, padx=20, font=12).place(x=0, y=430)


def openFiles():
    window.filename = filedialog.askopenfilename(initialdir="/", title="Select file",
                                                 filetypes=(("Excel Files", "*.xlsx"), ("all files", "*.*")))
    filename = " \n" + window.filename + " \n"
    tkinter.Label(window, text=filename, font=12).place(x=20, y=240)
    buttonVal2.set(0)
    return filename

# Splitting the file path for opening files
def splitter():
    filename = window.filename
    filename = filename.split('/')
    print(filename)
    print(len(filename))
    path = '//'.join(filename[0:len(filename)-1])
    print(path)
    filename = filename[len(filename)-1]
    print(filename)
    return path, filename

# ____________MAIN SPREADSHEET FUNCTION_____________ #

def create():
    if window.filename is not 'Empty':
        tkinter.Label(window, text="File Saved             ", justify=tkinter.CENTER, padx=20, font=12).place(x=185, y=500)
        if buttonVal.get() == 0:
            print('Fees and Enrolments')
            [file_path, input_file_name] = splitter()
            chdir(file_path)
            wb = openpyxl.load_workbook(input_file_name)

            sheet = wb.worksheets[0]
            counter = 2
            counter2 = 2
            total = 0
            matching = 0

            #  FUNDING AMOUNT
            sheet['AM1'] = "First Name"
            sheet['AN1'] = "Last Name"
            sheet['AO1'] = "Funding Amount"
            sheet['AP1'] = "Total"

            while counter <= sheet.max_row-1:

                if sheet['AE' + str(counter)].value != 0:
                    sheet['AM' + str(counter2)] = sheet['B' + str(counter)].value
                    sheet['AN' + str(counter2)] = sheet['C' + str(counter)].value
                    sheet['AO' + str(counter2)] = sheet['AE' + str(counter)].value
                    counter2 = counter2 + 1
                counter = counter + 1

            #  TOTAL CALCULATIONS
            counter = 2
            while counter < counter2:
                name1 = str(sheet['AM' + str(counter)].value) + ' ' + str(sheet['AN' + str(counter)].value)
                name2 = str(sheet['AM' + str(counter + 1)].value) + ' ' + str(sheet['AN' + str(counter + 1)].value)

                if (name1 == name2) & (matching == 0):
                    total = sheet['AO' + str(counter)].value + sheet['AO' + str(counter + 1)].value
                    matching = 1
                elif (name1 == name2) & (matching == 1):
                    total = total + sheet['AO' + str(counter + 1)].value

                if name1 != name2:
                    if matching == 0:
                        total = sheet['AO' + str(counter)].value
                    sheet['AP' + str(counter)] = total
                    matching = 0
                    total = 0
                counter = counter + 1

            #  PAID UPFRONT
            sheet['AR' + str(counter2)] = "First Name"
            sheet['AS' + str(counter2)] = "Last Name"
            sheet['AT' + str(counter2)] = "Unit Enrolment Status"
            sheet['AU' + str(counter2)] = "Paid Upfront"
            sheet['AV' + str(counter2)] = "Total"

            counter2 = counter2 + 1
            maximum = counter2
            counter = 2

            while counter <= sheet.max_row-1:

                if sheet['AD' + str(counter)].value != 0:
                    sheet['AR' + str(counter2)] = sheet['B' + str(counter)].value
                    sheet['AS' + str(counter2)] = sheet['C' + str(counter)].value
                    sheet['AT' + str(counter2)] = sheet['U' + str(counter)].value
                    sheet['AU' + str(counter2)] = sheet['AD' + str(counter)].value
                    counter2 = counter2 + 1
                counter = counter + 1

            #  TOTAL CALCULATION
            counter = maximum
            matching = 0
            total = 0

            while counter < counter2:
                name1 = str(sheet['AR' + str(counter)].value) + ' ' + str(sheet['AS' + str(counter)].value)
                name2 = str(sheet['AR' + str(counter + 1)].value) + ' ' + str(sheet['AS' + str(counter + 1)].value)
                if (name1 == name2) & (matching == 0):
                    total = sheet['AU' + str(counter)].value + sheet['AU' + str(counter + 1)].value
                    matching = 1
                elif (name1 == name2) & (matching == 1):
                    total = total + sheet['AU' + str(counter + 1)].value

                if name1 != name2:
                    if matching == 0:
                        total = sheet['AU' + str(counter)].value
                    sheet['AV' + str(counter)] = total
                    matching = 0
                    total = 0
                counter = counter + 1

            for idx in range(2, maximum - 1):
                if sheet['AP' + str(idx)].value == None:
                    sheet.row_dimensions[idx].hidden = True

            for idx in range(maximum, sheet.max_row):
                if sheet['AV' + str(idx)].value == None:
                    sheet.row_dimensions[idx].hidden = True

            wb.save(window.directory + '/' + outputName.get() + '.xlsx')

        elif buttonVal.get() == 1:
            print('Donations')
            # SPLIT THE NAME INTO PATH AND NAME
            [file_path, input_file_name] = splitter()
            chdir(file_path)
            wb = openpyxl.load_workbook(input_file_name)

            sheet = wb.worksheets[0]
            counter = 13
            total = 0
            matching = 0

            while counter < sheet.max_row - 2:

                name1 = sheet['F' + str(counter)].value
                name2 = sheet['F' + str(counter + 1)].value

                if (name1 == name2) & (matching == 0):
                    total = sheet['H' + str(counter)].value + sheet['H' + str(counter + 1)].value
                    matching = 1
                elif (name1 == name2) & (matching == 1):
                    total = total + sheet['H' + str(counter + 1)].value

                if name1 != name2:
                    if matching == 0:
                        total = sheet['H' + str(counter)].value
                    sheet['J' + str(counter)] = name1
                    sheet['K' + str(counter)] = total
                    matching = 0
                    total = 0

                counter = counter + 1

            for idx in range(13, sheet.max_row - 2):
                if sheet['K' + str(idx)].value == None:
                    sheet.row_dimensions[idx].hidden = True

            wb.save(window.directory + '/' + outputName.get() + '.xlsx')

    else:
        tkinter.Label(window, text="Select a file!", justify=tkinter.CENTER, padx=20, font=12).place(x=185, y=500)

# ____________TKINTER BUTTONS_____________ #

tkinter.Label(window, text="\nANONYMISED Spreadsheet Manager\n\n", justify=tkinter.CENTER, padx=20
              , font=16).pack(side=tkinter.TOP)

tkinter.Label(window, text="Select a spreadsheet type:", justify=tkinter.CENTER, padx=20, font=12).pack(side=tkinter.TOP)

for val, spreadsheet in enumerate(spreadsheets):
    tkinter.Radiobutton(window, text=spreadsheet[0], indicatoron=1, width=20, padx=20, variable=buttonVal, command=showChoice
                        , value=val, font=12).pack(side=tkinter.TOP)

file = tkinter.Button(window, text='Open File', width=25, command=showChoice2, font=12).place(x=200, y=200)

tkinter.Label(window, text='Output file name: ', font=20).place(x=20, y=320)

outputName = tkinter.StringVar()

entry = tkinter.Entry(window, textvariable=outputName, width=25, bd=2, font=22).place(x=200, y=320)

directoryButton = tkinter.Button(window, text='Output Folder', width=25, command=output2, font=12).place(x=200, y=370)

createButton = tkinter.Button(window, text='Create', width=25, command=create, font=12).place(x=200, y=475)

exitButton = tkinter.Button(window, text='Exit', width=25, command=window.destroy, font=12).place(x=350, y=550)


window.mainloop()

print(outputName.get())
print(window.filename)
print(window.directory)
