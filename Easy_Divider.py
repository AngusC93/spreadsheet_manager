#  EXCEL DONOR SUMMARY REPORT #
#  ANGUS CLARKE JULY 2019     #

import openpyxl
import os, time


print("Example path: C://Users//Directory_with_xlsx_file")
file_path = input("Enter file path: ")

print("Do not add file extension")
input_file_name = input("Enter Spreadsheet Name: ") + ".xlsx"

print("Do not add file extension")
output_file_name = input("Enter Output Name: ") + ".xlsx"



# ____________DO NOT EDIT PAST THIS LINE_____________ #

os.chdir(file_path)
wb = openpyxl.load_workbook(input_file_name)

sheet = wb["Sheet1"]
counter = 13
total = 0
matching = 0

while counter < sheet.max_row-2:

    name1 = sheet['F' + str(counter)].value
    name2 = sheet['F' + str(counter+1)].value

    if (name1 == name2) & (matching == 0):
        total = sheet['H' + str(counter)].value + sheet['H' + str(counter + 1)].value
        matching = 1
    elif (name1 == name2) & (matching == 1):
        total = total + sheet['H' + str(counter + 1)].value

    if name1 != name2:
        if matching == 0:
            total = sheet['H' + str(counter)].value
        sheet['J' + str(counter)] = name1
        sheet['K' + str(counter)] = total
        matching = 0
        total = 0

    counter = counter + 1

for idx in range(13, sheet.max_row-2):
    if sheet['K' + str(idx)].value == None:
        sheet.row_dimensions[idx].hidden = True

print("Completed. Have a nice day.")
wb.save(output_file_name)
time.sleep(5)
